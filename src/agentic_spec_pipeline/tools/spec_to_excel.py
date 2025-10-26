#!/usr/bin/env python3
"""
Deterministic spec.json to Excel converter.
No LLM required - pure transformation of structured JSON to human-readable Excel format.

Usage:
    python spec_to_excel.py <spec.json> [output.xlsx]

Example:
    python spec_to_excel.py examples/technical_requirements/spec.json review.xlsx
"""

import json
import sys
from pathlib import Path
from typing import Any, Dict, List
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class SpecToExcelConverter:
    """Converts spec.json to formatted Excel workbook for business review."""

    # Color scheme for visual clarity
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    SUBHEADER_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    SUBHEADER_FONT = Font(bold=True, size=10)
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def __init__(self, spec_path: str):
        """Initialize converter with spec file path."""
        self.spec_path = Path(spec_path)
        with open(self.spec_path, 'r') as f:
            self.spec = json.load(f)
        self.wb = Workbook()
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])

    def convert(self, output_path: str) -> None:
        """Main conversion method - creates all sheets."""
        self._create_overview_sheet()
        self._create_schema_sheet()

        # Create a sheet for each model
        for idx, model in enumerate(self.spec.get('models', []), 1):
            self._create_model_sheet(model, idx)

        # Create summary sheets
        self._create_all_columns_sheet()
        self._create_data_lineage_sheet()

        # Save workbook
        self.wb.save(output_path)
        print(f"✓ Excel file created: {output_path}")

    def _create_overview_sheet(self) -> None:
        """Create overview/summary sheet."""
        ws = self.wb.create_sheet("Overview", 0)

        row = 1
        # Title
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws[f'A{row}']
        cell.value = "Data Pipeline Specification - Overview"
        cell.font = Font(size=16, bold=True, color="366092")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 2

        # Metadata section
        ws[f'A{row}'] = "Source File:"
        ws[f'B{row}'] = str(self.spec_path.name)
        ws[f'A{row}'].font = Font(bold=True)
        row += 2

        # Models summary
        ws[f'A{row}'] = "Models Summary"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1

        # Headers
        headers = ["Model Name", "Layer", "Source Tables", "Target Columns", "Has Joins", "Has Aggregations"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER
        row += 1

        # Model rows
        for model in self.spec.get('models', []):
            ws.cell(row, 1, model.get('name', ''))
            ws.cell(row, 2, model.get('layer', ''))
            ws.cell(row, 3, ', '.join(model.get('sources', [])))
            ws.cell(row, 4, len(model.get('column_mapping', [])) + len(model.get('aggregations', []) or []))
            ws.cell(row, 5, 'Yes' if model.get('joins') else 'No')
            ws.cell(row, 6, 'Yes' if model.get('aggregations') else 'No')

            for col in range(1, 7):
                ws.cell(row, col).border = self.BORDER
                ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
            row += 1

        # Auto-size columns
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20

        ws.column_dimensions['C'].width = 40

    def _create_schema_sheet(self) -> None:
        """Create schema configuration sheet."""
        ws = self.wb.create_sheet("Schema Config")

        row = 1
        ws[f'A{row}'] = "Schema Configuration"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 2

        schema_config = self.spec.get('schema', {})

        headers = ["Configuration", "Value"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.border = self.BORDER
        row += 1

        for key, value in schema_config.items():
            ws.cell(row, 1, key)
            ws.cell(row, 2, value)
            for col in range(1, 3):
                ws.cell(row, col).border = self.BORDER
            row += 1

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25

    def _create_model_sheet(self, model: Dict[str, Any], idx: int) -> None:
        """Create detailed sheet for each model."""
        model_name = model.get('name', f'model_{idx}')
        ws = self.wb.create_sheet(f"{idx}. {model_name[:25]}")

        row = 1

        # Model header
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws[f'A{row}']
        cell.value = f"Model: {model_name}"
        cell.font = Font(size=14, bold=True, color="366092")
        cell.alignment = Alignment(horizontal='center')
        row += 2

        # Basic info
        ws[f'A{row}'] = "Layer:"
        ws[f'B{row}'] = model.get('layer', '')
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        ws[f'A{row}'] = "Source Tables:"
        ws[f'B{row}'] = ', '.join(model.get('sources', []))
        ws[f'A{row}'].font = Font(bold=True)
        row += 2

        # Column Mappings
        if model.get('column_mapping'):
            ws[f'A{row}'] = "Column Mappings"
            ws[f'A{row}'].font = self.SUBHEADER_FONT
            ws[f'A{row}'].fill = self.SUBHEADER_FILL
            row += 1

            headers = ["Target Column", "Type", "From Table", "From Column", "Transform",
                      "Nullable", "Tests", "Description"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row, col, header)
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
                cell.border = self.BORDER
            row += 1

            for col_map in model['column_mapping']:
                ws.cell(row, 1, col_map.get('target_column', ''))
                ws.cell(row, 2, col_map.get('type', ''))
                ws.cell(row, 3, col_map.get('from_table', ''))
                ws.cell(row, 4, col_map.get('from_column', ''))
                ws.cell(row, 5, col_map.get('transform', '') or '')
                ws.cell(row, 6, 'Yes' if col_map.get('nullable') else 'No')
                ws.cell(row, 7, ', '.join(col_map.get('tests', [])))
                ws.cell(row, 8, col_map.get('description', '') or '')

                for col in range(1, 9):
                    cell = ws.cell(row, col)
                    cell.border = self.BORDER
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                row += 1
            row += 1

        # Aggregations
        if model.get('aggregations'):
            ws[f'A{row}'] = "Aggregations"
            ws[f'A{row}'].font = self.SUBHEADER_FONT
            ws[f'A{row}'].fill = self.SUBHEADER_FILL
            row += 1

            headers = ["Metric Column", "Type", "Formula", "Tests", "Description"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row, col, header)
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.border = self.BORDER
            row += 1

            for agg in model['aggregations']:
                ws.cell(row, 1, agg.get('metric_column', ''))
                ws.cell(row, 2, agg.get('type', ''))
                ws.cell(row, 3, agg.get('formula', ''))
                ws.cell(row, 4, ', '.join(agg.get('tests', [])))
                ws.cell(row, 5, agg.get('description', '') or '')

                for col in range(1, 6):
                    cell = ws.cell(row, col)
                    cell.border = self.BORDER
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                row += 1
            row += 1

        # Group By
        if model.get('group_by'):
            ws[f'A{row}'] = "Group By:"
            ws[f'B{row}'] = ', '.join(model['group_by'])
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        # Joins
        if model.get('joins'):
            row += 1
            ws[f'A{row}'] = "Joins"
            ws[f'A{row}'].font = self.SUBHEADER_FONT
            ws[f'A{row}'].fill = self.SUBHEADER_FILL
            row += 1

            headers = ["Left Table", "Right Table", "Join Type", "Join Condition"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row, col, header)
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.border = self.BORDER
            row += 1

            for join in model['joins']:
                ws.cell(row, 1, join.get('left_table', ''))
                ws.cell(row, 2, join.get('right_table', ''))
                ws.cell(row, 3, join.get('type', ''))
                ws.cell(row, 4, join.get('condition', ''))

                for col in range(1, 5):
                    ws.cell(row, col).border = self.BORDER
                row += 1
            row += 1

        # Filters
        if model.get('filters'):
            ws[f'A{row}'] = "Filters"
            ws[f'A{row}'].font = self.SUBHEADER_FONT
            ws[f'A{row}'].fill = self.SUBHEADER_FILL
            row += 1

            headers = ["Applies To", "Predicate", "Rationale"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row, col, header)
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.border = self.BORDER
            row += 1

            for filt in model['filters']:
                ws.cell(row, 1, filt.get('applies_to', ''))
                ws.cell(row, 2, filt.get('predicate', ''))
                ws.cell(row, 3, filt.get('rationale', ''))

                for col in range(1, 4):
                    cell = ws.cell(row, col)
                    cell.border = self.BORDER
                    cell.alignment = Alignment(horizontal='left', wrap_text=True)
                row += 1
            row += 1

        # Constraints
        if model.get('constraints'):
            row += 1
            ws[f'A{row}'] = "Constraints"
            ws[f'A{row}'].font = self.SUBHEADER_FONT
            ws[f'A{row}'].fill = self.SUBHEADER_FILL
            row += 1

            for key, value in model['constraints'].items():
                ws.cell(row, 1, key)
                ws.cell(row, 2, str(value) if value else '')
                ws.cell(row, 1).font = Font(bold=True)
                ws.cell(row, 1).border = self.BORDER
                ws.cell(row, 2).border = self.BORDER
                row += 1

        # Auto-size columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 35
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 30

    def _create_all_columns_sheet(self) -> None:
        """Create a consolidated view of all columns across all models."""
        ws = self.wb.create_sheet("All Columns")

        row = 1
        ws[f'A{row}'] = "All Columns Across Models"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 2

        headers = ["Model", "Layer", "Column Name", "Type", "Source", "Transform",
                   "Nullable", "Tests", "Description"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.BORDER
        row += 1

        for model in self.spec.get('models', []):
            model_name = model.get('name', '')
            layer = model.get('layer', '')

            # Column mappings
            for col_map in model.get('column_mapping', []):
                ws.cell(row, 1, model_name)
                ws.cell(row, 2, layer)
                ws.cell(row, 3, col_map.get('target_column', ''))
                ws.cell(row, 4, col_map.get('type', ''))

                from_table = col_map.get('from_table', '')
                from_col = col_map.get('from_column', '')
                source = f"{from_table}.{from_col}" if from_table and from_col else ''
                ws.cell(row, 5, source)

                ws.cell(row, 6, col_map.get('transform', '') or '')
                ws.cell(row, 7, 'Yes' if col_map.get('nullable') else 'No')
                ws.cell(row, 8, ', '.join(col_map.get('tests', [])))
                ws.cell(row, 9, col_map.get('description', '') or '')

                for col in range(1, 10):
                    cell = ws.cell(row, col)
                    cell.border = self.BORDER
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                row += 1

            # Aggregations
            for agg in model.get('aggregations', []) or []:
                ws.cell(row, 1, model_name)
                ws.cell(row, 2, layer)
                ws.cell(row, 3, agg.get('metric_column', ''))
                ws.cell(row, 4, agg.get('type', ''))
                ws.cell(row, 5, 'AGGREGATION')
                ws.cell(row, 6, agg.get('formula', ''))
                ws.cell(row, 7, '')
                ws.cell(row, 8, ', '.join(agg.get('tests', [])))
                ws.cell(row, 9, agg.get('description', '') or '')

                for col in range(1, 10):
                    cell = ws.cell(row, col)
                    cell.border = self.BORDER
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                row += 1

        # Auto-size
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 30

    def _create_data_lineage_sheet(self) -> None:
        """Create data lineage/dependency view."""
        ws = self.wb.create_sheet("Data Lineage")

        row = 1
        ws[f'A{row}'] = "Data Lineage & Dependencies"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 2

        headers = ["Model", "Layer", "Depends On (Sources)", "Used By (Downstream)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.BORDER
        row += 1

        # Build dependency map
        model_names = {m['name'] for m in self.spec.get('models', [])}
        downstream_map = {name: [] for name in model_names}

        for model in self.spec.get('models', []):
            for source in model.get('sources', []):
                # Extract model name from source (e.g., "stg_customers" from sources)
                for potential_upstream in model_names:
                    if potential_upstream in source:
                        downstream_map[potential_upstream].append(model['name'])

        # Write lineage
        for model in self.spec.get('models', []):
            model_name = model.get('name', '')
            layer = model.get('layer', '')
            sources = ', '.join(model.get('sources', []))
            downstream = ', '.join(sorted(set(downstream_map.get(model_name, []))))

            ws.cell(row, 1, model_name)
            ws.cell(row, 2, layer)
            ws.cell(row, 3, sources)
            ws.cell(row, 4, downstream)

            for col in range(1, 5):
                cell = ws.cell(row, col)
                cell.border = self.BORDER
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            row += 1

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 40


def main():
    """CLI entry point."""
    if len(sys.argv) < 2:
        print("Usage: python spec_to_excel.py <spec.json> [output.xlsx]")
        print("\nExample:")
        print("  python spec_to_excel.py examples/technical_requirements/spec.json")
        print("  python spec_to_excel.py examples/technical_requirements/spec.json custom_output.xlsx")
        sys.exit(1)

    spec_path = sys.argv[1]

    if not Path(spec_path).exists():
        print(f"Error: File not found: {spec_path}")
        sys.exit(1)

    # Default output to examples/human_review/ directory
    if len(sys.argv) >= 3:
        output_path = sys.argv[2]
    else:
        # Use examples/human_review/ if the spec is in examples/
        spec_file = Path(spec_path)
        if 'examples' in spec_file.parts:
            output_dir = Path('examples/human_review')
            output_dir.mkdir(parents=True, exist_ok=True)
            output_path = output_dir / f"{spec_file.stem}_review.xlsx"
        else:
            output_path = Path(spec_path).stem + "_review.xlsx"

    try:
        converter = SpecToExcelConverter(spec_path)
        converter.convert(output_path)
        print(f"\n✓ Successfully converted {spec_path}")
        print(f"✓ Output: {output_path}")
        print(f"\nThe Excel file contains:")
        print(f"  • Overview sheet with model summary")
        print(f"  • Schema configuration")
        print(f"  • Detailed sheet for each model")
        print(f"  • All columns consolidated view")
        print(f"  • Data lineage & dependencies")
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()